# ============================================================
# FINTRACK — app.py
# ============================================================


from flask import (Flask, render_template, request,
                   redirect, url_for, flash, send_file)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (LoginManager, UserMixin,
                         login_user, logout_user,
                         login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from translations import get_translation
import secrets
import os
import re
import csv
import io


from PIL import Image
import pytesseract


pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'




# ============================================================
# APP CONFIG
# ============================================================


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI']      = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY']                   = 'fintrack-secret'
app.config['UPLOAD_FOLDER']                = 'static/uploads'
app.config['MAX_CONTENT_LENGTH']           = 2 * 1024 * 1024  # 2 MB


os.makedirs('static/uploads', exist_ok=True)


db           = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view    = 'login'
login_manager.login_message = 'Silakan login dulu!'




# ============================================================
# MODELS
# ============================================================


class User(UserMixin, db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    name         = db.Column(db.String(100), nullable=False)
    email        = db.Column(db.String(100), unique=True, nullable=False)
    password     = db.Column(db.String(200), nullable=False)
    photo        = db.Column(db.String(200), nullable=True)
    language     = db.Column(db.String(5),  default='id')
    theme        = db.Column(db.String(10), default='light')
    notif_popup  = db.Column(db.Boolean,    default=True)
    timezone     = db.Column(db.String(50), default='Asia/Jakarta')
    created_at   = db.Column(db.DateTime,   default=datetime.utcnow)
    transactions = db.relationship('Transaction', backref='user', lazy=True)




class LoginHistory(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    user_id      = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    ip_address   = db.Column(db.String(50),  nullable=True)
    user_agent   = db.Column(db.String(200), nullable=True)
    logged_in_at = db.Column(db.DateTime, default=datetime.utcnow)
    user         = db.relationship('User', backref='login_histories')




class Transaction(db.Model):
    id       = db.Column(db.Integer, primary_key=True)
    title    = db.Column(db.String(100), nullable=False)
    amount   = db.Column(db.Float,       nullable=False)
    category = db.Column(db.String(50),  nullable=False)
    type     = db.Column(db.String(10),  nullable=False)
    date     = db.Column(db.DateTime,    default=datetime.utcnow)
    note     = db.Column(db.String(200), nullable=True)
    user_id  = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)




class Budget(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    amount     = db.Column(db.Float,      nullable=False)
    period     = db.Column(db.String(10), nullable=False)  # daily / weekly / monthly
    created_at = db.Column(db.DateTime,   default=datetime.utcnow)




class Group(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100), nullable=False)
    code       = db.Column(db.String(8),   unique=True, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    members    = db.relationship('GroupMember',     backref='group', lazy=True)
    funds      = db.relationship('GroupFund',       backref='group', lazy=True)
    logs       = db.relationship('ActivityLog',     backref='group', lazy=True)




class GroupMember(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    group_id  = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=False)
    user_id   = db.Column(db.Integer, db.ForeignKey('user.id'),  nullable=False)
    role      = db.Column(db.String(10), default='member')  # admin / member
    joined_at = db.Column(db.DateTime,  default=datetime.utcnow)
    user      = db.relationship('User', backref='group_memberships')




class GroupFund(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    group_id    = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=False)
    name        = db.Column(db.String(100), nullable=False)
    target      = db.Column(db.Float,       default=0)
    description = db.Column(db.String(200), nullable=True)
    created_at  = db.Column(db.DateTime,    default=datetime.utcnow)




class GroupTransaction(db.Model):
    id       = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'),      nullable=False)
    fund_id  = db.Column(db.Integer, db.ForeignKey('group_fund.id'), nullable=True)
    user_id  = db.Column(db.Integer, db.ForeignKey('user.id'),       nullable=False)
    title    = db.Column(db.String(100), nullable=False)
    amount   = db.Column(db.Float,       nullable=False)
    category = db.Column(db.String(50),  nullable=False)
    type     = db.Column(db.String(10),  nullable=False)
    date     = db.Column(db.DateTime,    default=datetime.utcnow)
    note     = db.Column(db.String(200), nullable=True)
    user     = db.relationship('User',      backref='group_transactions')
    fund     = db.relationship('GroupFund', backref='transactions')




class ActivityLog(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    group_id   = db.Column(db.Integer, db.ForeignKey('group.id'),  nullable=False)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'),   nullable=False)
    action     = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime,    default=datetime.utcnow)
    user       = db.relationship('User', backref='activity_logs')




class SplitBill(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    group_id     = db.Column(db.Integer, db.ForeignKey('group.id'),  nullable=False)
    created_by   = db.Column(db.Integer, db.ForeignKey('user.id'),   nullable=False)
    title        = db.Column(db.String(100), nullable=False)
    total_amount = db.Column(db.Float,       nullable=False)
    note         = db.Column(db.String(200), nullable=True)
    created_at   = db.Column(db.DateTime,    default=datetime.utcnow)
    splits       = db.relationship('SplitDetail', backref='bill', lazy=True)
    creator      = db.relationship('User', backref='split_bills')




class SplitDetail(db.Model):
    id      = db.Column(db.Integer, primary_key=True)
    bill_id = db.Column(db.Integer, db.ForeignKey('split_bill.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'),       nullable=False)
    amount  = db.Column(db.Float,   nullable=False)
    is_paid = db.Column(db.Boolean, default=False)
    paid_at = db.Column(db.DateTime, nullable=True)
    user    = db.relationship('User', backref='split_details')




# ============================================================
# HELPERS
# ============================================================


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))




@app.context_processor
def inject_translation():
    lang = current_user.language if current_user.is_authenticated else 'id'
    return dict(t=get_translation(lang))




def log_activity(group_id, user_id, action):
    db.session.add(ActivityLog(group_id=group_id, user_id=user_id, action=action))
    db.session.commit()




def get_budget_status(user_id):
    budget = Budget.query.filter_by(user_id=user_id).first()
    if not budget:
        return None


    now = datetime.utcnow()


    if budget.period == 'daily':
        start        = now.replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Hari Ini'
    elif budget.period == 'weekly':
        start        = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Minggu Ini'
    else:
        start        = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Bulan Ini'


    transactions  = Transaction.query.filter(
        Transaction.user_id == user_id,
        Transaction.type    == 'expense',
        Transaction.date    >= start
    ).all()


    total_expense = sum(t.amount for t in transactions)
    percentage    = (total_expense / budget.amount * 100) if budget.amount > 0 else 0
    remaining     = budget.amount - total_expense


    if percentage >= 100:
        status  = 'danger'
        message = f'Budget {period_label} sudah habis! Pengeluaran melebihi Rp {budget.amount:,.0f}'
    elif percentage >= 90:
        status  = 'warning'
        message = f'Hampir habis! Sisa budget {period_label} tinggal Rp {remaining:,.0f}'
    else:
        status  = 'safe'
        message = None


    return {
        'amount'       : budget.amount,
        'period'       : budget.period,
        'period_label' : period_label,
        'total_expense': total_expense,
        'remaining'    : remaining,
        'percentage'   : min(percentage, 100),
        'status'       : status,
        'message'      : message,
    }




def _ocr_receipt(file_stream):
    """Baca struk via Tesseract, kembalikan (title, amount, note)."""
    img   = Image.open(file_stream)
    text  = pytesseract.image_to_string(img, lang='ind+eng')
    lines = text.split('\n')


    amount = 0
    for line in lines:
        if any(w in line.lower() for w in ['total', 'grand total', 'jumlah', 'bayar']):
            for num in reversed(re.findall(r'[\d.,]+', line)):
                clean = num.replace('.', '').replace(',', '')
                if clean.isdigit() and int(clean) > 1000:
                    amount = int(clean)
                    break
            if amount:
                break


    title = 'Belanja'
    for line in lines[:5]:
        line = line.strip()
        if len(line) > 3 and not line.isdigit():
            title = line.title()
            break


    note = ', '.join(l.strip() for l in lines if len(l.strip()) > 3)[:200]
    return title, amount, note




# ============================================================
# LANDING PAGE
# ============================================================


@app.route('/')
def landing():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    return render_template('landing.html')




# ============================================================
# AUTH
# ============================================================


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))


    if request.method == 'POST':
        name     = request.form['name']
        email    = request.form['email']
        password = request.form['password']
        confirm  = request.form['confirm']


        if password != confirm:
            flash('Password tidak cocok!', 'danger')
            return redirect(url_for('register'))


        if User.query.filter_by(email=email).first():
            flash('Email sudah terdaftar!', 'danger')
            return redirect(url_for('register'))


        new_user = User(name=name, email=email,
                        password=generate_password_hash(password))
        db.session.add(new_user)
        db.session.commit()


        flash('Akun berhasil dibuat! Silakan login.', 'success')
        return redirect(url_for('login'))


    return render_template('register.html')




@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))


    if request.method == 'POST':
        email    = request.form['email']
        password = request.form['password']
        user     = User.query.filter_by(email=email).first()


        if not user or not check_password_hash(user.password, password):
            flash('Email atau password salah!', 'danger')
            return redirect(url_for('login'))


        login_user(user)


        # Catat riwayat login
        db.session.add(LoginHistory(
            user_id    = user.id,
            ip_address = request.remote_addr,
            user_agent = request.headers.get('User-Agent', '')[:200]
        ))
        db.session.commit()


        flash(f'Selamat datang, {user.name}!', 'success')
        return redirect(url_for('index'))


    return render_template('login.html')




@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Berhasil logout!', 'success')
    return redirect(url_for('login'))




# ============================================================
# MAIN — TRANSAKSI PERSONAL
# ============================================================


@app.route('/index')
@login_required
def index():
    transactions  = Transaction.query.filter_by(user_id=current_user.id)\
                               .order_by(Transaction.date.desc()).all()
    total_income  = sum(t.amount for t in transactions if t.type == 'income')
    total_expense = sum(t.amount for t in transactions if t.type == 'expense')
    balance       = total_income - total_expense
    budget_status = get_budget_status(current_user.id)
    budget        = Budget.query.filter_by(user_id=current_user.id).first()


    return render_template('index.html',
                           transactions=transactions,
                           total_income=total_income,
                           total_expense=total_expense,
                           balance=balance,
                           budget_status=budget_status,
                           budget=budget)




@app.route('/add', methods=['GET', 'POST'])
@login_required
def add_transaction():
    if request.method == 'POST':
        date_str = request.form.get('date', '')
        tanggal  = datetime.strptime(date_str, '%Y-%m-%d') if date_str else datetime.utcnow()


        db.session.add(Transaction(
            title    = request.form['title'],
            amount   = float(request.form['amount']),
            category = request.form['category'],
            type     = request.form['type'],
            note     = request.form.get('note', ''),
            date     = tanggal,
            user_id  = current_user.id
        ))
        db.session.commit()
        flash('Transaksi berhasil ditambahkan!', 'success')
        return redirect(url_for('index'))


    today = datetime.utcnow().strftime('%Y-%m-%d')
    return render_template('add_transaction.html', today=today)




@app.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_transaction(id):
    transaction = Transaction.query.filter_by(id=id, user_id=current_user.id).first_or_404()


    if request.method == 'POST':
        transaction.title    = request.form['title']
        transaction.amount   = float(request.form['amount'])
        transaction.category = request.form['category']
        transaction.type     = request.form['type']
        transaction.note     = request.form.get('note', '')
        transaction.date     = datetime.strptime(request.form['date'], '%Y-%m-%d')
        db.session.commit()
        flash('Transaksi berhasil diperbarui!', 'success')
        return redirect(url_for('index'))


    return render_template('edit_transaction.html', transaction=transaction)




@app.route('/delete/<int:id>')
@login_required
def delete_transaction(id):
    transaction = Transaction.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    db.session.delete(transaction)
    db.session.commit()
    flash('Transaksi berhasil dihapus!', 'danger')
    return redirect(url_for('index'))




# ============================================================
# LAPORAN
# ============================================================


@app.route('/report')
@login_required
def report():
    transactions  = Transaction.query.filter_by(user_id=current_user.id)\
                               .order_by(Transaction.date.asc()).all()


    category_data = {}
    monthly_data  = {}


    for t in transactions:
        if t.type == 'expense':
            category_data[t.category] = category_data.get(t.category, 0) + t.amount
        key = t.date.strftime('%b %Y')
        if key not in monthly_data:
            monthly_data[key] = {'income': 0, 'expense': 0}
        monthly_data[key][t.type] += t.amount


    total_income  = sum(t.amount for t in transactions if t.type == 'income')
    total_expense = sum(t.amount for t in transactions if t.type == 'expense')
    balance       = total_income - total_expense


    return render_template('report.html',
                           category_data=category_data,
                           monthly_data=monthly_data,
                           total_income=total_income,
                           total_expense=total_expense,
                           balance=balance)




# ============================================================
# SCAN STRUK
# ============================================================


@app.route('/scan', methods=['GET', 'POST'])
@login_required
def scan():
    if request.method == 'POST':
        file = request.files.get('receipt')
        if not file or file.filename == '':
            flash('Pilih file dulu!', 'danger')
            return redirect(url_for('scan'))


        title, amount, note = _ocr_receipt(file.stream)
        result = {
            'title'   : title,
            'amount'  : amount,
            'category': 'Belanja',
            'date'    : datetime.utcnow().strftime('%Y-%m-%d'),
            'note'    : note,
            'user_id' : current_user.id
        }
        return render_template('scan.html', result=result)


    return render_template('scan.html', result=None)




# ============================================================
# BUDGET
# ============================================================


@app.route('/budget', methods=['POST'])
@login_required
def budget():
    amount   = float(request.form['amount'])
    period   = request.form['period']
    existing = Budget.query.filter_by(user_id=current_user.id).first()


    if existing:
        existing.amount = amount
        existing.period = period
    else:
        db.session.add(Budget(user_id=current_user.id, amount=amount, period=period))


    db.session.commit()
    flash('Budget berhasil disimpan!', 'success')
    return redirect(url_for('index'))




# ============================================================
# GRUP — SHARING MODE
# ============================================================


@app.route('/groups')
@login_required
def groups():
    memberships = GroupMember.query.filter_by(user_id=current_user.id).all()
    return render_template('groups.html', groups=[m.group for m in memberships])




@app.route('/groups/create', methods=['POST'])
@login_required
def create_group():
    name      = request.form['name']
    code      = secrets.token_hex(4).upper()
    new_group = Group(name=name, code=code, created_by=current_user.id)
    db.session.add(new_group)
    db.session.flush()
    db.session.add(GroupMember(group_id=new_group.id, user_id=current_user.id, role='admin'))
    db.session.commit()
    log_activity(new_group.id, current_user.id, f'membuat grup "{name}"')
    flash(f'Grup "{name}" berhasil dibuat! Kode: {code}', 'success')
    return redirect(url_for('group_detail', group_id=new_group.id))




@app.route('/groups/join', methods=['POST'])
@login_required
def join_group():
    code  = request.form['code'].strip().upper()
    group = Group.query.filter_by(code=code).first()


    if not group:
        flash('Kode grup tidak ditemukan!', 'danger')
        return redirect(url_for('groups'))


    if GroupMember.query.filter_by(group_id=group.id, user_id=current_user.id).first():
        flash('Kamu sudah bergabung di grup ini!', 'warning')
        return redirect(url_for('group_detail', group_id=group.id))


    db.session.add(GroupMember(group_id=group.id, user_id=current_user.id, role='member'))
    db.session.commit()
    log_activity(group.id, current_user.id, 'bergabung ke grup')
    flash(f'Berhasil bergabung ke grup "{group.name}"!', 'success')
    return redirect(url_for('group_detail', group_id=group.id))




@app.route('/groups/<int:group_id>')
@login_required
def group_detail(group_id):
    group  = Group.query.get_or_404(group_id)
    member = GroupMember.query.filter_by(group_id=group_id, user_id=current_user.id).first()
    if not member:
        flash('Kamu bukan anggota grup ini!', 'danger')
        return redirect(url_for('groups'))


    transactions  = GroupTransaction.query.filter_by(group_id=group_id)\
                                    .order_by(GroupTransaction.date.desc()).all()
    funds         = GroupFund.query.filter_by(group_id=group_id).all()
    logs          = ActivityLog.query.filter_by(group_id=group_id)\
                               .order_by(ActivityLog.created_at.desc()).limit(20).all()
    total_income  = sum(t.amount for t in transactions if t.type == 'income')
    total_expense = sum(t.amount for t in transactions if t.type == 'expense')


    return render_template('group_detail.html',
                           group=group, member=member,
                           transactions=transactions, funds=funds, logs=logs,
                           total_income=total_income, total_expense=total_expense,
                           balance=total_income - total_expense,
                           now=datetime.utcnow())


@app.route('/groups/<int:group_id>/scan', methods=['GET', 'POST'])
@login_required
def scan_group_transaction(group_id):
    group  = Group.query.get_or_404(group_id)
    member = GroupMember.query.filter_by(
        group_id=group_id, user_id=current_user.id).first()
    if not member:
        flash('Kamu bukan anggota grup ini!', 'danger')
        return redirect(url_for('groups'))

    funds = GroupFund.query.filter_by(group_id=group_id).all()

    if request.method == 'POST':
        file = request.files.get('receipt')
        if not file or file.filename == '':
            flash('Pilih file dulu!', 'danger')
            return redirect(url_for('scan_group_transaction', group_id=group_id))

        title, amount, note = _ocr_receipt(file.stream)
        result = {
            'title'   : title,
            'amount'  : amount,
            'category': 'Belanja',
            'date'    : datetime.utcnow().strftime('%Y-%m-%d'),
            'note'    : note,
            'user_id' : current_user.id
        }
        return render_template('scan_group_transaction.html',
                               group=group, funds=funds, result=result)

    return render_template('scan_group_transaction.html',
                           group=group, funds=funds, result=None)


@app.route('/groups/<int:group_id>/add', methods=['GET', 'POST'])
@login_required
def add_group_transaction(group_id):
    group = Group.query.get_or_404(group_id)

    if request.method == 'POST':
        mode = request.form.get('mode')

        if mode == 'scan':
            file = request.files.get('receipt')
            if not file or file.filename == '':
                flash('Pilih file struk dulu!', 'danger')
                return redirect(url_for('add_group_transaction', group_id=group_id))

            items, total = _ocr_receipt(file.stream)  # fungsi OCR kamu
            transaction = GroupTransaction(
                group_id=group.id,
                title="Scan Struk",
                category="Belanja",
                type="expense",
                amount=total,
                date=datetime.utcnow(),
                user_id  = current_user.id, 
            )
            db.session.add(transaction)
            db.session.commit()
            flash('Transaksi dari scan berhasil ditambahkan!', 'success')
            return redirect(url_for('group_detail', group_id=group.id))

        else:  # manual
            transaction = GroupTransaction(
                group_id=group.id,
                title=request.form['title'],
                category=request.form['category'],
                type=request.form['type'],
                amount=float(request.form['amount']),
                date=datetime.utcnow(),
                user_id  = current_user.id
            )
            db.session.add(transaction)
            db.session.commit()
            flash('Transaksi manual berhasil ditambahkan!', 'success')
            return redirect(url_for('group_detail', group_id=group.id))

    return render_template('add_group_transaction.html', group=group)





@app.route('/groups/<int:group_id>/delete/<int:trans_id>')
@login_required
def delete_group_transaction(group_id, trans_id):
    t = GroupTransaction.query.filter_by(id=trans_id, group_id=group_id).first_or_404()
    log_activity(group_id, current_user.id, f'menghapus transaksi "{t.title}" Rp {t.amount:,.0f}')
    db.session.delete(t)
    db.session.commit()
    flash('Transaksi berhasil dihapus!', 'danger')
    return redirect(url_for('group_detail', group_id=group_id))




@app.route('/groups/<int:group_id>/edit/<int:trans_id>', methods=['GET', 'POST'])
@login_required
def edit_group_transaction(group_id, trans_id):
    group  = Group.query.get_or_404(group_id)
    member = GroupMember.query.filter_by(group_id=group_id, user_id=current_user.id).first()
    if not member:
        flash('Kamu bukan anggota grup ini!', 'danger')
        return redirect(url_for('groups'))


    t = GroupTransaction.query.filter_by(id=trans_id, group_id=group_id).first_or_404()


    if request.method == 'POST':
        old_title, old_amount = t.title, t.amount
        t.title    = request.form['title']
        t.amount   = float(request.form['amount'])
        t.category = request.form['category']
        t.type     = request.form['type']
        t.note     = request.form.get('note', '')
        t.fund_id  = request.form.get('fund_id') or None
        date_str   = request.form.get('date', '')
        if date_str:
            t.date = datetime.strptime(date_str, '%Y-%m-%d')
        db.session.commit()
        log_activity(group_id, current_user.id,
                     f'mengedit transaksi "{old_title}" (Rp {old_amount:,.0f} → Rp {t.amount:,.0f})')
        flash('Transaksi berhasil diperbarui!', 'success')
        return redirect(url_for('group_detail', group_id=group_id))


    funds = GroupFund.query.filter_by(group_id=group_id).all()
    return render_template('edit_group_transaction.html', group=group, transaction=t, funds=funds)




# ── Dana / Pos Keuangan ──


@app.route('/groups/<int:group_id>/fund/add', methods=['POST'])
@login_required
def add_fund(group_id):
    name = request.form['fund_name']
    db.session.add(GroupFund(
        group_id    = group_id,
        name        = name,
        target      = float(request.form.get('fund_target', 0)),
        description = request.form.get('fund_desc', '')
    ))
    db.session.commit()
    log_activity(group_id, current_user.id, f'membuat dana "{name}"')
    flash(f'Dana "{name}" berhasil dibuat!', 'success')
    return redirect(url_for('group_detail', group_id=group_id))




@app.route('/groups/<int:group_id>/fund/edit/<int:fund_id>', methods=['POST'])
@login_required
def edit_fund(group_id, fund_id):
    fund = GroupFund.query.filter_by(id=fund_id, group_id=group_id).first_or_404()
    if not GroupMember.query.filter_by(group_id=group_id, user_id=current_user.id).first():
        flash('Kamu bukan anggota grup ini!', 'danger')
        return redirect(url_for('groups'))


    old_name         = fund.name
    fund.name        = request.form['fund_name']
    fund.target      = float(request.form.get('fund_target', 0))
    fund.description = request.form.get('fund_desc', '')
    db.session.commit()
    log_activity(group_id, current_user.id, f'mengedit dana "{old_name}" menjadi "{fund.name}"')
    flash(f'Dana "{fund.name}" berhasil diperbarui!', 'success')
    return redirect(url_for('group_detail', group_id=group_id))




@app.route('/groups/<int:group_id>/fund/delete/<int:fund_id>')
@login_required
def delete_fund(group_id, fund_id):
    fund = GroupFund.query.filter_by(id=fund_id, group_id=group_id).first_or_404()
    if not GroupMember.query.filter_by(group_id=group_id, user_id=current_user.id).first():
        flash('Kamu bukan anggota grup ini!', 'danger')
        return redirect(url_for('groups'))


    log_activity(group_id, current_user.id, f'menghapus dana "{fund.name}"')
    db.session.delete(fund)
    db.session.commit()
    flash(f'Dana "{fund.name}" berhasil dihapus!', 'danger')
    return redirect(url_for('group_detail', group_id=group_id))




# ── Laporan Grup ──


@app.route('/groups/<int:group_id>/report')
@login_required
def group_report(group_id):
    group  = Group.query.get_or_404(group_id)
    member = GroupMember.query.filter_by(group_id=group_id, user_id=current_user.id).first()
    if not member:
        flash('Kamu bukan anggota grup ini!', 'danger')
        return redirect(url_for('groups'))


    transactions  = GroupTransaction.query.filter_by(group_id=group_id)\
                                    .order_by(GroupTransaction.date.asc()).all()


    category_data = {}
    monthly_data  = {}
    member_data   = {}
    fund_data     = {}


    for t in transactions:
        key = t.date.strftime('%b %Y')
        if key not in monthly_data:
            monthly_data[key] = {'income': 0, 'expense': 0}
        monthly_data[key][t.type] += t.amount


        if t.type == 'expense':
            category_data[t.category]           = category_data.get(t.category, 0) + t.amount
            member_data[t.user.name]             = member_data.get(t.user.name, 0) + t.amount
            fname                                = t.fund.name if t.fund else 'Tanpa Dana'
            fund_data[fname]                     = fund_data.get(fname, 0) + t.amount


    total_income  = sum(t.amount for t in transactions if t.type == 'income')
    total_expense = sum(t.amount for t in transactions if t.type == 'expense')


    return render_template('group_report.html',
                           group=group,
                           category_data=category_data, monthly_data=monthly_data,
                           member_data=member_data, fund_data=fund_data,
                           total_income=total_income, total_expense=total_expense,
                           balance=total_income - total_expense)




# ── Split Tagihan ──


@app.route('/groups/<int:group_id>/split')
@login_required
def split_list(group_id):
    group  = Group.query.get_or_404(group_id)
    member = GroupMember.query.filter_by(group_id=group_id, user_id=current_user.id).first()
    if not member:
        flash('Kamu bukan anggota grup ini!', 'danger')
        return redirect(url_for('groups'))


    bills   = SplitBill.query.filter_by(group_id=group_id)\
                       .order_by(SplitBill.created_at.desc()).all()
    members = GroupMember.query.filter_by(group_id=group_id).all()
    return render_template('split_bill.html', group=group, bills=bills,
                           members=members, current_user=current_user)




@app.route('/groups/<int:group_id>/split/create', methods=['POST'])
@login_required
def create_split(group_id):
    if not GroupMember.query.filter_by(group_id=group_id, user_id=current_user.id).first():
        flash('Kamu bukan anggota grup ini!', 'danger')
        return redirect(url_for('groups'))


    member_ids   = request.form.getlist('member_ids')
    if not member_ids:
        flash('Pilih minimal satu anggota!', 'danger')
        return redirect(url_for('split_list', group_id=group_id))


    total_amount = float(request.form['total_amount'])
    title        = request.form['title']
    split_type   = request.form['split_type']


    bill = SplitBill(group_id=group_id, created_by=current_user.id,
                     title=title, total_amount=total_amount,
                     note=request.form.get('note', ''))
    db.session.add(bill)
    db.session.flush()


    for uid in member_ids:
        amount = (total_amount / len(member_ids)) if split_type == 'equal' \
                 else float(request.form.get(f'custom_{uid}', 0))
        db.session.add(SplitDetail(bill_id=bill.id, user_id=int(uid), amount=amount))


    db.session.commit()
    log_activity(group_id, current_user.id,
                 f'membuat split tagihan "{title}" total Rp {total_amount:,.0f} untuk {len(member_ids)} orang')
    flash('Split tagihan berhasil dibuat!', 'success')
    return redirect(url_for('split_list', group_id=group_id))




@app.route('/groups/<int:group_id>/split/<int:bill_id>/pay/<int:detail_id>')
@login_required
def mark_paid(group_id, bill_id, detail_id):
    detail         = SplitDetail.query.get_or_404(detail_id)
    detail.is_paid = True
    detail.paid_at = datetime.utcnow()
    db.session.commit()
    log_activity(group_id, current_user.id,
                 f'menandai pembayaran "{detail.bill.title}" Rp {detail.amount:,.0f} sebagai lunas')
    flash('Pembayaran berhasil ditandai lunas!', 'success')
    return redirect(url_for('split_list', group_id=group_id))




@app.route('/groups/<int:group_id>/split/<int:bill_id>/delete')
@login_required
def delete_split(group_id, bill_id):
    bill = SplitBill.query.filter_by(id=bill_id, group_id=group_id).first_or_404()
    if bill.created_by != current_user.id:
        flash('Hanya pembuat tagihan yang bisa menghapus!', 'danger')
        return redirect(url_for('split_list', group_id=group_id))


    log_activity(group_id, current_user.id, f'menghapus split tagihan "{bill.title}"')
    db.session.delete(bill)
    db.session.commit()
    flash('Split tagihan berhasil dihapus!', 'danger')
    return redirect(url_for('split_list', group_id=group_id))




@app.route('/groups/<int:group_id>/split/scan', methods=['GET', 'POST'])
@login_required
def scan_split(group_id):
    group  = Group.query.get_or_404(group_id)
    member = GroupMember.query.filter_by(group_id=group_id, user_id=current_user.id).first()
    if not member:
        flash('Kamu bukan anggota grup ini!', 'danger')
        return redirect(url_for('groups'))


    members = GroupMember.query.filter_by(group_id=group_id).all()


    if request.method == 'POST':
        file = request.files.get('receipt')
        if not file or file.filename == '':
            flash('Pilih file dulu!', 'danger')
            return redirect(url_for('scan_split', group_id=group_id))


        title, amount, note = _ocr_receipt(file.stream)
        result = {'title': title, 'amount': amount, 'note': note}
        return render_template('scan_split.html', group=group, members=members, result=result)


    return render_template('scan_split.html', group=group, members=members, result=None)




# ============================================================
# SETTINGS
# ============================================================


@app.route('/settings')
@login_required
def settings():
    login_histories = LoginHistory.query.filter_by(user_id=current_user.id)\
                                  .order_by(LoginHistory.logged_in_at.desc()).limit(5).all()
    return render_template('settings.html', login_histories=login_histories)

@app.route('/settings/profile', methods=['POST'])
@login_required
def update_profile():
    new_email = request.form['email']
    existing  = User.query.filter_by(email=new_email).first()
    if existing and existing.id != current_user.id:
        flash('Email sudah dipakai akun lain!', 'danger')
        return redirect(url_for('settings'))


    current_user.name  = request.form['name']
    current_user.email = new_email


    file = request.files.get('photo')
    if file and file.filename:
        ext = file.filename.rsplit('.', 1)[-1].lower()
        if ext in ['jpg', 'jpeg', 'png', 'gif']:
            filename = secure_filename(f'user_{current_user.id}.{ext}')
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            current_user.photo = filename
        else:
            flash('Format foto tidak didukung! Gunakan JPG/PNG.', 'danger')
            return redirect(url_for('settings'))


    db.session.commit()
    flash('Profil berhasil diperbarui!', 'success')
    return redirect(url_for('settings'))




@app.route('/settings/password', methods=['POST'])
@login_required
def change_password():
    current_pw = request.form['current_password']
    new_pw     = request.form['new_password']
    confirm_pw = request.form['confirm_password']


    if not check_password_hash(current_user.password, current_pw):
        flash('Password saat ini salah!', 'danger')
    elif new_pw != confirm_pw:
        flash('Password baru tidak cocok!', 'danger')
    elif len(new_pw) < 6:
        flash('Password minimal 6 karakter!', 'danger')
    else:
        current_user.password = generate_password_hash(new_pw)
        db.session.commit()
        flash('Password berhasil diubah!', 'success')


    return redirect(url_for('settings'))




@app.route('/settings/preferences', methods=['POST'])
@login_required
def update_preferences():
    current_user.language    = request.form.get('language', 'id')
    current_user.theme       = request.form.get('theme', 'light')
    current_user.notif_popup = 'notif_popup' in request.form
    current_user.timezone    = request.form.get('timezone', 'Asia/Jakarta')
    db.session.commit()
    flash('Preferensi berhasil disimpan!', 'success')
    return redirect(url_for('settings'))




@app.route('/settings/export/csv')
@login_required
def export_csv():
    transactions = Transaction.query.filter_by(user_id=current_user.id)\
                              .order_by(Transaction.date.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Tanggal', 'Judul', 'Kategori', 'Tipe', 'Jumlah', 'Catatan'])
    for t in transactions:
        writer.writerow([
            t.date.strftime('%d/%m/%Y'), t.title, t.category,
            'Pemasukan' if t.type == 'income' else 'Pengeluaran',
            t.amount, t.note or ''
        ])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8-sig')),
                     mimetype='text/csv', as_attachment=True,
                     download_name=f'fintrack_{current_user.name}_transaksi.csv')




@app.route('/settings/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment


    transactions = Transaction.query.filter_by(user_id=current_user.id)\
                              .order_by(Transaction.date.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transaksi'


    headers = ['Tanggal', 'Judul', 'Kategori', 'Tipe', 'Jumlah (Rp)', 'Catatan']
    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='2D6A4F')
        cell.alignment = Alignment(horizontal='center')


    for row, t in enumerate(transactions, 2):
        vals = [t.date.strftime('%d/%m/%Y'), t.title, t.category,
                'Pemasukan' if t.type == 'income' else 'Pengeluaran',
                t.amount, t.note or '']
        fill_color = 'D1FAE5' if t.type == 'income' else 'FEE2E2'
        for col, val in enumerate(vals, 1):
            cell      = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill('solid', fgColor=fill_color)


    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = \
            min(max(len(str(c.value or '')) for c in col) + 4, 40)


    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'fintrack_{current_user.name}_transaksi.xlsx')




@app.route('/settings/delete-account', methods=['POST'])
@login_required
def delete_account():
    if not check_password_hash(current_user.password, request.form['password']):
        flash('Password salah! Akun tidak dihapus.', 'danger')
        return redirect(url_for('settings'))


    Transaction.query.filter_by(user_id=current_user.id).delete()
    Budget.query.filter_by(user_id=current_user.id).delete()
    LoginHistory.query.filter_by(user_id=current_user.id).delete()
    GroupMember.query.filter_by(user_id=current_user.id).delete()


    user = User.query.get(current_user.id)
    logout_user()
    db.session.delete(user)
    db.session.commit()


    flash('Akun berhasil dihapus.', 'success')
    return redirect(url_for('login'))




# ============================================================
# RUN
# ============================================================


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        print("Database siap!")
    app.run(debug=True)